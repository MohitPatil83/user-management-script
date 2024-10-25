import openpyxl
import os




def add_user(file_name):
    workbook = create_or_load_excel(file_name)
    sheet = workbook.active

    name = input("Enter Name: ")
    email = input("Enter Email: ")
    phone = input("Enter Phone Number: ")

    sheet.append([name, email, phone])
    workbook.save(file_name)
    print(f"User '{name}' added successfully.\n")


def display_users(file_name):
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active

        if sheet.max_row > 1:
            print(f"{'Name':<20} {'Email':<30} {'Phone Number':<15}")
            print("-" * 65)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                print(f"{row[0]:<20} {row[1]:<30} {row[2]:<15}")
        else:
            print("No users found.\n")
    else:
        print("No data available. Please add users first.\n")

def create_or_load_excel(file_name):
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        sheet.append(["Name", "Email", "Phone Number"])
        workbook.save(file_name)
    else:
        workbook = openpyxl.load_workbook(file_name)
    return workbook


def main():
    file_name = "users.xlsx"

    while True:
        print("1. Add User")
        print("2. Display Users")
        print("3. Exit")
        choice = input("Choose an option: ")

        if choice == "1":
            add_user(file_name)
        elif choice == "2":
            display_users(file_name)
        elif choice == "3":
            print("Exiting the program.")
            break
        else:
            print("Invalid choice. Please try again.\n")

if __name__ == "__main__":
    main()
